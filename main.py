from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Lista de números de telefone a serem alterados
numeros_telefone = ['999999787', '999999528','999999246', '999998856'] 
caminho_arquivo_original = 'C:\\Users\\lucas\\PROJETOS_PYTHON\\alteracaoLucas.xlsx'
caminho_arquivo_modificado = 'C:\\Users\\lucas\\PROJETOS_PYTHON\\testeExcel\\NovoTeste.xlsx'

# Carrega o workbook 
wb = load_workbook(caminho_arquivo_original)
ws = wb['Distrito Federal']  # Seleciona a aba "Distrito Federal"

# Definir a cor vermelha para o preenchimento
fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Seleciona a coluna telefone e itera sobre a planilha para achá-la
coluna_telefone_idx = None
for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1))):
    if cell.value == 'Telefone':
        coluna_telefone_idx = col_idx
        break

if coluna_telefone_idx is None:
    print("A coluna 'Telefone' não foi encontrada.")
else:
    # Iterar pelas linhas e verificar se o telefone está na lista
    for row in ws.iter_rows(min_row=2):  # Começar da segunda linha para pular o cabeçalho
        telefone = str(row[coluna_telefone_idx].value).strip()
        if telefone in numeros_telefone:
            # Aplicar a cor vermelha na linha completa
            for cell in row:
                cell.fill = fill_red

    # Salvar o arquivo modificado
    wb.save(caminho_arquivo_modificado)

    print("Planilha modificada criada com sucesso.")
